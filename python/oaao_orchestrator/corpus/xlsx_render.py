"""Table / blocks → XLSX (CS-3-S4)."""

from __future__ import annotations

import io
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from oaao_orchestrator.library.blocks import blocks_to_markdown


def _parse_markdown_table(md: str) -> tuple[list[str], list[list[str]]]:
    lines = [ln.strip() for ln in (md or "").splitlines() if ln.strip()]
    for i, ln in enumerate(lines):
        if not ln.startswith("|") or i + 1 >= len(lines):
            continue
        sep = lines[i + 1]
        if not re.match(r"^\|?\s*:?-+:?\s*(\|\s*:?-+:?\s*)+\|?\s*$", sep):
            continue
        header_cells = [c.strip() for c in ln.strip("|").split("|")]
        rows: list[list[str]] = []
        for body_ln in lines[i + 2 :]:
            if not body_ln.startswith("|"):
                break
            rows.append([c.strip() for c in body_ln.strip("|").split("|")])
        return header_cells, rows
    return [], []


def table_data_to_xlsx_bytes(
    *,
    columns: list[tuple[str, str]] | list[str],
    rows: list[dict[str, Any]] | list[list[str]],
    title: str = "",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = (title or "Sheet1")[:31] or "Sheet1"

    headers: list[str] = []
    keys: list[str] = []
    if columns and isinstance(columns[0], tuple):
        for key, label in columns:  # type: ignore[index]
            keys.append(str(key))
            headers.append(str(label))
    elif columns:
        headers = [str(c) for c in columns]
        keys = headers[:]

    if headers:
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

    for row in rows:
        if isinstance(row, dict):
            ws.append([str(row.get(k) or "") for k in keys] if keys else list(row.values()))
        elif isinstance(row, list):
            ws.append([str(c) for c in row])

    for idx, _ in enumerate(headers or [""], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def markdown_to_xlsx_bytes(markdown: str, *, title: str = "") -> bytes:
    headers, rows = _parse_markdown_table(markdown)
    if headers:
        return table_data_to_xlsx_bytes(columns=headers, rows=rows, title=title)
    wb = Workbook()
    ws = wb.active
    ws.title = (title or "Sheet1")[:31] or "Sheet1"
    if title.strip():
        ws.append([title.strip()])
        ws["A1"].font = Font(bold=True)
    for raw_line in (markdown or "").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or line.startswith("```"):
            continue
        ws.append([line])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def blocks_to_xlsx_bytes(blocks: list[dict[str, Any]], *, title: str = "") -> bytes:
    for block in blocks:
        if not isinstance(block, dict):
            continue
        if str(block.get("type") or "").strip().lower() != "table":
            continue
        cols_raw = block.get("columns")
        rows_raw = block.get("rows")
        columns: list[tuple[str, str]] = []
        if isinstance(cols_raw, list):
            for c in cols_raw:
                if isinstance(c, dict) and c.get("key"):
                    columns.append((str(c["key"]), str(c.get("label") or c["key"])))
                elif isinstance(c, str):
                    columns.append((c, c))
        row_list: list[dict[str, Any]] = []
        if isinstance(rows_raw, list):
            for r in rows_raw:
                if isinstance(r, dict):
                    row_list.append(r)
        if columns and row_list:
            return table_data_to_xlsx_bytes(columns=columns, rows=row_list, title=title)
    md = blocks_to_markdown(blocks, title=title)
    return markdown_to_xlsx_bytes(md, title=title)
